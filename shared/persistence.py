from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def _abrir(ruta: Path, hoja: str) -> tuple[Workbook, Any]:
    if ruta.exists():
        wb = load_workbook(ruta)
        if hoja in wb.sheetnames:
            return wb, wb[hoja]
        ws = wb.create_sheet(hoja)
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    return wb, ws


def _tiene_encabezados(wb: Workbook, hoja: str) -> bool:
    ws = wb[hoja]
    if ws.max_row < 1:
        return False
    return any(c.value is not None for c in list(ws[1]))


def _escribir_encabezados(ws: Any, encabezados: list[str]) -> None:
    for i, h in enumerate(encabezados, start=1):
        ws.cell(row=1, column=i, value=h)


def leer_hoja(ruta_archivo: str | Path, hoja: str | None = None) -> list[dict[str, Any]]:
    ruta = Path(ruta_archivo)
    if not ruta.exists():
        return []
    wb = load_workbook(ruta, data_only=True)
    ws = wb[hoja] if hoja else wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        wb.close()
        return []
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(filas[0])]
    wb.close()
    return [dict(zip(headers, fila)) for fila in filas[1:]]


def escribir_fila(ruta_archivo: str | Path, hoja: str, datos: dict[str, Any]) -> None:
    ruta = Path(ruta_archivo)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb, ws = _abrir(ruta, hoja)

    if not _tiene_encabezados(wb, hoja):
        _escribir_encabezados(ws, list(datos.keys()))

    headers = [c.value for c in ws[1]]
    ws.append([datos.get(h, "") for h in headers])
    wb.save(ruta)
    wb.close()


def buscar_por_id(
    ruta_archivo: str | Path, hoja: str, id_columna: str, id_valor: Any
) -> dict[str, Any] | None:
    ruta = Path(ruta_archivo)
    if not ruta.exists():
        return None
    wb = load_workbook(ruta, data_only=True)
    ws = wb[hoja]
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        wb.close()
        return None
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(filas[0])]
    if id_columna not in headers:
        wb.close()
        return None
    idx = headers.index(id_columna)
    for fila in filas[1:]:
        if fila[idx] == id_valor:
            wb.close()
            return dict(zip(headers, fila))
    wb.close()
    return None


def actualizar(
    ruta_archivo: str | Path,
    hoja: str,
    id_columna: str,
    id_valor: Any,
    datos: dict[str, Any],
) -> bool:
    ruta = Path(ruta_archivo)
    if not ruta.exists():
        return False
    wb = load_workbook(ruta)
    ws = wb[hoja]
    headers = [str(c.value) if c.value is not None else f"col_{i}" for i, c in enumerate(ws[1])]
    if id_columna not in headers:
        wb.close()
        return False
    idx_id = headers.index(id_columna)
    for fila in ws.iter_rows(min_row=2):
        if fila[idx_id].value == id_valor:
            for col, val in datos.items():
                if col in headers:
                    fila[headers.index(col)].value = val
            wb.save(ruta)
            wb.close()
            return True
    wb.close()
    return False
